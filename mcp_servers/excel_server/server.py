import asyncio
import json

from mcp.server import Server
from mcp.types import Tool, TextContent, ServerCapabilities
from mcp.server.models import InitializationOptions
import os
import openpyxl
import openpyxl.cell
from openpyxl.utils import column_index_from_string
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from typing import Any, Dict, List

class ExcelMCPServer:
  def __init__(self):
    self.server = Server("excel-operations")
    self._setup_handlers()

  def _setup_handlers(self):

    @self.server.list_tools()
    async def list_tools() -> List[Tool]:
      return [
        Tool(
            name="read_range",
            description="Read values from an Excel range",
            inputSchema={
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Path to Excel file"
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "Name of worksheet",
                        "default": "Sheet1"
                    },
                    "range": {
                        "type": "string",
                        "description": "Excel range like 'A1:C10'"
                    }
                },
                "required": ["file_path", "range"]
            }
        ),
        Tool(
            name="write_range",
            description="Write values to an Excel range",
            inputSchema={
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Path to Excel file"
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "Name of worksheet",
                        "default": "Sheet1"
                    },
                    "range": {
                        "type": "string",
                        "description": "Excel range like 'A1:C10'"
                    },
                    "values": {
                        "type": "array",
                        "description": "2D array of values to write"
                    }
                },
                "required": ["file_path", "range", "values"]
            }
        ),
        Tool(
            name="get_workbook_info",
            description="Get basic information about a workbook",
            inputSchema={
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Path to Excel file"
                    }
                },
                "required": ["file_path"]
            }
        )
    ]

    @self.server.call_tool()
    async def call_tool(name: str, arguments: Dict[str, Any]) -> List[TextContent]:
      # route tool calls to the correct handlers
      try:
        if name == "read_range":
          return await self._read_range(arguments)
        elif name == "write_range":
          return await self._write_range(arguments)
        elif name == "get_workbook_info":
          return await self._get_workbook_info(arguments)
        else:
          return [TextContent(
            type="text",
            text=json.dumps({
              "success": False,
              "error": f"Unknown tool: {name}"
            })
          )]
      except Exception as e:
        # always catch teh exceptions at the top level
        return [TextContent(
            type="text",
            text=json.dumps({
            "success": False,
            "error": f'Tool execution failed: {str(e)}'
          })
        )]

  async def _read_range(self, args: Dict[str, Any]) -> List[TextContent]:
    # excel range reading

    try:
      # step 1: extract and validate arguments
      file_path = args["file_path"]
      sheet_name = args.get("sheet_name", "Sheet1")
      range_str = args["range"]

      # step 2: verify that the file exists
      if not os.path.exists(file_path):
        return [TextContent(
                    type="text",
                    text=json.dumps({
                        "success": False,
                        "error": f"File not found: {file_path}"
                    })
                )]

      # step 3: open the workbook
      workbook = openpyxl.load_workbook(file_path, data_only=True)

      # step 4: validate that the sheets exists
      if sheet_name not in workbook.sheetnames:
        return [TextContent(
                    type="text",
                    text=json.dumps({
                        "success": False,
                        "error": f"Sheet '{sheet_name}' not found. Available: {workbook.sheetnames}"
                    })
                )]
      worksheet = workbook[sheet_name]

      # step 5: simple, bulletproof approach
      try:
        if ":" in range_str:
          # Parse range coordinates
          start_cell, end_cell = range_str.split(":")
          start_col_letter, start_row = coordinate_from_string(start_cell)
          end_col_letter, end_row = coordinate_from_string(end_cell)
          start_col = column_index_from_string(start_col_letter)
          end_col = column_index_from_string(end_col_letter)

          values = []
          for row in range(start_row, end_row + 1):
            row_values = []
            for col in range(start_col, end_col + 1):
              cell_value = worksheet.cell(row=row, column=col).value
              row_values.append(cell_value if cell_value is not None else "")
            values.append(row_values)
        else:
          # Single cell - use direct cell access
          col_letter, row_num = coordinate_from_string(range_str)
          col_idx = column_index_from_string(col_letter)
          cell_value = worksheet.cell(row=row_num, column=col_idx).value
          values = [[cell_value if cell_value is not None else ""]]

      except Exception as e:
        values = [[f"Error: {str(e)}"]]

      # Step 7: Return structured success response
      return [TextContent(
        type="text",
        text=json.dumps({
          "success": True,
          "data": values,
          "metadata": {
            "file_path": file_path,
            "sheet_name": sheet_name,
            "range": range_str,
            "rows": len(values),
            "columns": len(values[0]) if values else 0
          }
        })
      )]

    except Exception as e:
      return [TextContent(
        type="text",
        text=json.dumps({
          "success": False,
          "error": f"Failed to read range: {str(e)}"
        })
      )]

  async def _write_range(self, args: Dict[str, Any]) -> List[TextContent]:
    try:
      # step 1: get arguments
      file_path = args["file_path"]
      sheet_name = args.get("sheet_name", "Sheet1")
      range_str = args["range"]
      values = args["values"]

      # step 2: make sure that the file path exists
      if not os.path.exists(file_path):
        return [TextContent(
          type="text",
          text=json.dumps({
            "success": False,
            "error": "File Path doesn't exists"
          })
        )]

      # step 3: create the workbook
      workbook = openpyxl.load_workbook(file_path, data_only=True)

      # step 4: verify that the worksheet exists
      if sheet_name not in workbook.sheetnames:
        return [TextContent(
          type="text",
          text=json.dumps({
            "success": False,
            "error": "Unable to find sheet name"
          })
        )]

      worksheet = workbook[sheet_name]

      # step 6: parse, and find out what from what row to col we need to write to
      # if range string looks like A1:B3
      start = range_str.split(":")[0]

      # covert to usable values for offset
      col_letter, start_row = coordinate_from_string(start)
      start_col = column_index_from_string(col_letter)

      # write each value in values to each cell
      for row_i, row in enumerate(values, start=start_row):
        for col_i, col in enumerate(row, start=start_col):
          # write to the row
          worksheet.cell(row=row_i - 1, column=col_i - 1, value = col)

      # step 7 save the notebook
      workbook.save(file_path)

      # success message
      return [TextContent(
        type='text',
        text=json.dumps({
          "success": True
        })
      )]

    except Exception as e:
      return [TextContent(
        type="text",
        text=json.dumps({
          "sucess": False,
          "error": "Failed to read excel file"
        })
      )]


  async def _get_workbook_info(self, args: Dict[str, Any]) -> List[TextContent]:
    # TODO: Implement workbook info functionality
    return [TextContent(
      type="text",
      text=json.dumps({
        "success": False,
        "error": "get_workbook_info not implemented yet"
      })
    )]

async def main():
  import sys
  from mcp.server.stdio import stdio_server

  server = ExcelMCPServer()
  print("Starting Excel MCP Server...", file=sys.stderr)

  # Use stdio_server for proper MCP communication
  async with stdio_server() as (read_stream, write_stream):
    init_options = InitializationOptions(
      server_name="excel-mcp-server",
      server_version="1.0.0",
      capabilities=ServerCapabilities()
    )
    await server.server.run(read_stream, write_stream, init_options)

if __name__ == "__main__":
  asyncio.run(main())
