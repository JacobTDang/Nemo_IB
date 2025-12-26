import openpyxl

# Simple test to understand openpyxl behavior
wb = openpyxl.Workbook()
ws = wb.active

# Add data
ws["A1"] = "Test"
ws["B1"] = 123
ws["A2"] = "Data" 
ws["B2"] = 456

wb.save("simple_test.xlsx")

# Now read it
wb = openpyxl.load_workbook("simple_test.xlsx", data_only=True)
ws = wb.active

print("=== Simple Tests ===")

# Test 1: Single cell
single = ws["A1"]
print(f"Single cell: {type(single)} = {single.value}")

# Test 2: Range
range_data = ws["A1:B2"]
print(f"Range type: {type(range_data)}")

# Test 3: What's in the range?
for i, row in enumerate(range_data):
    print(f"Row {i}: {type(row)}")
    if isinstance(row, tuple):
        for j, cell in enumerate(row):
            print(f"  Cell {j}: {type(cell)} = {getattr(cell, 'value', 'no value attr')}")
    else:
        print(f"  Not a tuple: {type(row)} = {getattr(row, 'value', 'no value attr')}")