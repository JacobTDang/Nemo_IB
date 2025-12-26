import openpyxl
import openpyxl.cell

# Test different cell range scenarios
def test_cell_ranges():
    # First create test data
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Add test data
    ws["A1"] = "Header1"
    ws["B1"] = "Header2" 
    ws["C1"] = "Header3"
    ws["A2"] = 100
    ws["B2"] = 200
    ws["C2"] = 300
    
    wb.save("debug_test.xlsx")
    
    # Now test different range types
    wb = openpyxl.load_workbook("debug_test.xlsx", data_only=True)
    ws = wb.active
    
    print("=== Testing Cell Range Types ===")
    
    # Test 1: Single cell
    single_cell = ws["A1"]
    print(f"Single cell A1: {type(single_cell)} = {single_cell.value}")
    
    # Test 2: Single row range
    single_row = ws["A1:C1"]
    print(f"Single row A1:C1: {type(single_row)}")
    print(f"First element type: {type(single_row[0])}")
    if hasattr(single_row[0], 'value'):
        print(f"First element has .value: {single_row[0].value}")
    else:
        print(f"First element IS the value: {single_row[0]}")
    
    # Test 3: Multi-row range
    multi_row = ws["A1:C2"]
    print(f"Multi-row A1:C2: {type(multi_row)}")
    print(f"First element type: {type(multi_row[0])}")
    print(f"First row, first cell type: {type(multi_row[0][0])}")
    if hasattr(multi_row[0][0], 'value'):
        print(f"First cell has .value: {multi_row[0][0].value}")
    else:
        print(f"First cell IS the value: {multi_row[0][0]}")
    
    # Test the iteration logic
    def extract_values(cell_range):
        if isinstance(cell_range, openpyxl.cell.Cell):
            return [[cell_range.value if cell_range.value is not None else ""]]
        else:
            values = []
            try:
                if isinstance(cell_range[0], openpyxl.cell.Cell):
                    # Single row
                    row_values = []
                    for cell in cell_range:
                        value = cell.value
                        row_values.append(value if value is not None else "")
                    values.append(row_values)
                else:
                    # Multi-row
                    for row in cell_range:
                        row_values = []
                        for cell in row:
                            value = cell.value
                            row_values.append(value if value is not None else "")
                        values.append(row_values)
            except Exception as e:
                print(f"Error: {e}")
                values = [["Error"]]
            return values
    
    print("\n=== Testing Value Extraction ===")
    print(f"Single cell: {extract_values(single_cell)}")
    print(f"Single row: {extract_values(single_row)}")
    print(f"Multi-row: {extract_values(multi_row)}")

if __name__ == "__main__":
    test_cell_ranges()